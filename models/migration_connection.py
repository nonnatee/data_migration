# -*- coding: utf-8 -*-

import base64
import csv
import io
import json
import logging
import os
import re
import struct
import time
import urllib.request
import urllib.error
import urllib.parse
from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)


class MigrationConnection(models.Model):
    _name = 'migration.connection'
    _description = 'Data Migration Connection'
    _order = 'name asc'

    name = fields.Char(string='Connection Name', required=True)
    conn_type = fields.Selection([
        ('file_csv', 'CSV / TSV File'),
        ('file_excel', 'Excel File (.xlsx / .xls)'),
        ('file_json', 'JSON File'),
        ('file_xml', 'XML File'),
        ('foxpro_dbf', 'Visual FoxPro DBF File / Directory'),
        ('odbc', 'ODBC Connection (DSN / Connection String)'),
        ('database_sql', 'Direct SQL Database (PostgreSQL / MySQL / MSSQL / SQLite / Oracle)'),
        ('api_rest', 'REST API (HTTP / JSON Endpoint)'),
        ('api_graphql', 'GraphQL Query Endpoint'),
        ('cloud_storage', 'Cloud Object Storage (AWS S3 / GCS / Azure / SFTP)'),
        ('google_sheets', 'Google Sheets (Public URL / CSV Export)'),
        ('odoo_rpc', 'Odoo XML-RPC / JSON-RPC Endpoint'),
    ], string='Connection Type', default='file_csv', required=True)

    # ------------------------------------------------------------
    # 1. FILE SOURCE CONFIGURATION
    # ------------------------------------------------------------
    source_type = fields.Selection([
        ('upload', 'File Upload'),
        ('path', 'Server File/Folder Path'),
        ('url', 'Remote HTTP/HTTPS URL'),
    ], string='Source Location', default='upload')
    file_binary = fields.Binary(string='Upload File', attachment=True)
    file_name = fields.Char(string='File Name')
    file_path = fields.Char(string='Server File / Directory Path', help='Absolute filesystem path on the Odoo server host.')
    file_url = fields.Char(string='Remote Download URL', placeholder='https://example.com/data/export.csv')
    file_encoding = fields.Selection([
        ('utf-8', 'UTF-8'),
        ('utf-8-sig', 'UTF-8 with BOM'),
        ('cp1252', 'Windows CP1252 (Western European)'),
        ('cp874', 'Windows CP874 / TIS-620 (Thai)'),
        ('cp850', 'DOS CP850 (Latin 1)'),
        ('latin1', 'ISO-8859-1 (Latin-1)'),
        ('gbk', 'GBK / GB2312 (Chinese)'),
        ('shift_jis', 'Shift-JIS (Japanese)'),
    ], string='File Encoding', default='utf-8')

    # CSV / TSV Options
    csv_delimiter = fields.Char(string='CSV Delimiter', default=',')
    csv_quotechar = fields.Char(string='CSV Quote Character', default='"')
    csv_has_header = fields.Boolean(string='First Row is Header', default=True)

    # Excel Options
    excel_sheet_name = fields.Char(string='Excel Sheet Name / Index', placeholder='Leave empty for active/first sheet, or specify sheet name')

    # XML / JSON Options
    json_data_path = fields.Char(string='JSON Data Key Path', placeholder='e.g. data.items or results', help='Path to array inside JSON response/file.')
    xml_root_xpath = fields.Char(string='XML Record XPath', default='.//record', placeholder='.//row or .//item or .//record')

    # FoxPro DBF Options
    dbf_table_name = fields.Char(string='DBF Table Name', help='Specify DBF file name if directory path is specified.')
    dbf_memo_path = fields.Char(string='Memo File Path (.fpt / .dbt)', help='Optional path to associated memo file for text fields.')

    # ------------------------------------------------------------
    # 2. ODBC CONFIGURATION
    # ------------------------------------------------------------
    odbc_connection_string = fields.Char(
        string='ODBC Connection String',
        help='Driver={FoxPro Files};Dbq=C:\\Data; or DSN=MyLegacyDSN;UID=user;PWD=pass;'
    )

    # ------------------------------------------------------------
    # 3. SQL DATABASE CONFIGURATION
    # ------------------------------------------------------------
    db_type = fields.Selection([
        ('postgresql', 'PostgreSQL'),
        ('mysql', 'MySQL / MariaDB'),
        ('mssql', 'Microsoft SQL Server'),
        ('sqlite', 'SQLite 3 (.db / .sqlite file)'),
        ('oracle', 'Oracle Database'),
    ], string='Database Type', default='postgresql')
    db_host = fields.Char(string='Host / Server', default='localhost')
    db_port = fields.Integer(string='Port', default=5432)
    db_name = fields.Char(string='Database Name / File Path')
    db_user = fields.Char(string='Username')
    db_password = fields.Char(string='Password')
    db_ssl_mode = fields.Selection([
        ('disable', 'Disable'),
        ('prefer', 'Prefer'),
        ('require', 'Require SSL'),
    ], string='SSL Mode', default='prefer')
    db_query = fields.Text(string='SQL Query / Table Name', default='SELECT * FROM source_table',
                           help='e.g., SELECT id, name, email FROM customers WHERE active = 1')

    # ------------------------------------------------------------
    # 4. REST API & GRAPHQL CONFIGURATION
    # ------------------------------------------------------------
    api_url = fields.Char(string='API Endpoint URL', placeholder='https://api.example.com/v1/customers')
    api_http_method = fields.Selection([
        ('GET', 'GET'),
        ('POST', 'POST'),
    ], string='HTTP Method', default='GET')
    api_auth_type = fields.Selection([
        ('none', 'No Auth'),
        ('bearer', 'Bearer Token'),
        ('api_key', 'API Key Header / Query Param'),
        ('basic', 'HTTP Basic Auth'),
        ('oauth2', 'OAuth 2.0 Client Credentials'),
    ], string='Authentication Type', default='none')
    api_token = fields.Char(string='API Token / Secret Key')
    api_key_header = fields.Char(string='API Key Header Name', default='X-API-Key')
    api_user = fields.Char(string='Auth Username / Client ID')
    api_password = fields.Char(string='Auth Password / Client Secret')
    api_headers_json = fields.Text(string='Custom Headers (JSON)', default='{"Accept": "application/json"}')
    api_params_json = fields.Text(string='Query Parameters (JSON)', default='{}')
    api_request_body = fields.Text(string='Request Body (JSON / GraphQL)', default='{}')
    api_pagination_type = fields.Selection([
        ('none', 'No Pagination (Single Request)'),
        ('page_number', 'Page Number (?page=1&per_page=100)'),
        ('offset_limit', 'Offset & Limit (?offset=0&limit=100)'),
        ('cursor_link', 'Next URL / Cursor Token'),
    ], string='Pagination Strategy', default='none')
    api_page_param = fields.Char(string='Page Param Name', default='page')
    api_limit_param = fields.Char(string='Limit / Per Page Param', default='limit')
    api_page_size = fields.Integer(string='Page Size', default=100)
    api_max_pages = fields.Integer(string='Max Pages Limit', default=50, help='Safety limit for automated pagination.')

    # ------------------------------------------------------------
    # 5. CLOUD STORAGE & GOOGLE SHEETS
    # ------------------------------------------------------------
    cloud_provider = fields.Selection([
        ('s3', 'Amazon Web Services (AWS S3)'),
        ('gcs', 'Google Cloud Storage (GCS)'),
        ('azure_blob', 'Microsoft Azure Blob Storage'),
        ('sftp', 'SFTP / FTP Server'),
    ], string='Cloud Storage Provider', default='s3')
    cloud_bucket = fields.Char(string='Bucket / Container Name')
    cloud_object_key = fields.Char(string='Object Key / Remote File Path', placeholder='exports/customers_2026.csv')
    cloud_access_key = fields.Char(string='Access Key ID / Account Name')
    cloud_secret_key = fields.Char(string='Secret Access Key / SAS Token')
    cloud_region = fields.Char(string='Region', default='us-east-1')
    cloud_endpoint_url = fields.Char(string='Custom S3 Endpoint URL', placeholder='https://s3.compat.example.com')
    
    # SFTP specific
    sftp_host = fields.Char(string='SFTP Host')
    sftp_port = fields.Integer(string='SFTP Port', default=22)
    sftp_user = fields.Char(string='SFTP Username')
    sftp_password = fields.Char(string='SFTP Password')

    # Google Sheets
    sheets_url = fields.Char(string='Google Sheet URL', placeholder='https://docs.google.com/spreadsheets/d/YOUR_SHEET_ID/edit')
    sheets_gid = fields.Char(string='Sheet GID / Tab Name', default='0')

    # ------------------------------------------------------------
    # 6. ODOO RPC CONFIGURATION
    # ------------------------------------------------------------
    rpc_url = fields.Char(string='Odoo URL', placeholder='https://odoo.example.com')
    rpc_db = fields.Char(string='Remote Database')
    rpc_user = fields.Char(string='Remote User')
    rpc_password = fields.Char(string='API Key / Password')
    rpc_model = fields.Char(string='Remote Model Name', placeholder='res.partner')
    rpc_domain = fields.Char(string='Search Domain Filter', default='[]')

    # ------------------------------------------------------------
    # STATUS, TIMING & SCHEMA METADATA
    # ------------------------------------------------------------
    state = fields.Selection([
        ('draft', 'Draft'),
        ('connected', 'Connected / Ready'),
        ('error', 'Connection Error'),
    ], string='Status', default='draft', readonly=True)
    last_error = fields.Text(string='Last Connection Error', readonly=True)
    latency_ms = fields.Float(string='Response Latency (ms)', readonly=True)
    source_columns = fields.Text(string='Discovered Source Columns (JSON)', help='JSON array of column names.')
    source_schema_info = fields.Text(string='Schema Types & Metadata (JSON)', help='JSON object with column metadata.')
    preview_data = fields.Text(string='Sample Data Preview (JSON)', help='JSON preview of top records.')

    # ------------------------------------------------------------
    # CONNECTION TESTING & DATA FETCHING
    # ------------------------------------------------------------

    def action_test_connection(self):
        """Test connection, measure latency, and extract schema columns + preview data."""
        self.ensure_one()
        start_ts = time.time()
        try:
            records, columns, schema_info = self._fetch_raw_records_with_schema(limit=10)
            duration_ms = round((time.time() - start_ts) * 1000.0, 2)
            self.write({
                'state': 'connected',
                'last_error': False,
                'latency_ms': duration_ms,
                'source_columns': json.dumps(columns),
                'source_schema_info': json.dumps(schema_info),
                'preview_data': json.dumps(records[:10], default=str),
            })
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Connection Successful'),
                    'message': _('Connected in %s ms. Discovered %d columns from %s.', duration_ms, len(columns), self.conn_type),
                    'type': 'success',
                    'sticky': False,
                }
            }
        except Exception as e:
            _logger.exception("Failed to test connection ID %s (%s)", self.id, self.name)
            self.write({
                'state': 'error',
                'last_error': str(e),
                'latency_ms': 0.0,
            })
            raise UserError(_("Connection failed: %s") % str(e))

    def _fetch_raw_records_with_schema(self, limit=None):
        """Fetches raw records and constructs schema info dictionary."""
        records, columns = self._fetch_raw_records(limit=limit)
        schema_info = {}
        if records and isinstance(records, list):
            for col in columns:
                sample_vals = [r.get(col) for r in records if r.get(col) is not None and str(r.get(col)).strip() != '']
                inferred_type = 'string'
                if sample_vals:
                    first = sample_vals[0]
                    if isinstance(first, bool):
                        inferred_type = 'boolean'
                    elif isinstance(first, int):
                        inferred_type = 'integer'
                    elif isinstance(first, float):
                        inferred_type = 'float'
                    elif isinstance(first, (dict, list)):
                        inferred_type = 'json'
                    elif re.match(r'^\d{4}-\d{2}-\d{2}', str(first)):
                        inferred_type = 'date'
                schema_info[col] = {
                    'inferred_type': inferred_type,
                    'sample_value': str(sample_vals[0]) if sample_vals else '',
                    'null_count': sum(1 for r in records if r.get(col) is None or str(r.get(col)).strip() == ''),
                }
        return records, columns, schema_info

    def _fetch_raw_records(self, limit=None):
        """Unified raw record fetcher returning (list_of_dicts, list_of_column_names)."""
        self.ensure_one()
        if self.conn_type == 'file_csv':
            return self._parse_csv(limit=limit)
        elif self.conn_type == 'file_excel':
            return self._parse_excel(limit=limit)
        elif self.conn_type == 'file_json':
            return self._parse_json(limit=limit)
        elif self.conn_type == 'file_xml':
            return self._parse_xml(limit=limit)
        elif self.conn_type == 'foxpro_dbf':
            return self._parse_foxpro_dbf(limit=limit)
        elif self.conn_type == 'odbc':
            return self._parse_odbc(limit=limit)
        elif self.conn_type == 'database_sql':
            return self._parse_sql_db(limit=limit)
        elif self.conn_type == 'api_rest':
            return self._parse_rest_api(limit=limit)
        elif self.conn_type == 'api_graphql':
            return self._parse_graphql(limit=limit)
        elif self.conn_type == 'cloud_storage':
            return self._parse_cloud_storage(limit=limit)
        elif self.conn_type == 'google_sheets':
            return self._parse_google_sheets(limit=limit)
        elif self.conn_type == 'odoo_rpc':
            return self._parse_odoo_rpc(limit=limit)
        else:
            raise UserError(_("Unsupported connection type: %s") % self.conn_type)

    # ------------------------------------------------------------
    # 1. FILE & STREAM RETRIEVAL HELPERS
    # ------------------------------------------------------------

    def _get_file_content_bytes(self):
        """Retrieve bytes from binary upload, local filesystem path, or remote URL."""
        if self.source_type == 'upload':
            if not self.file_binary:
                raise UserError(_("Please upload a file."))
            return base64.b64decode(self.file_binary)
        elif self.source_type == 'path':
            if not self.file_path or not os.path.exists(self.file_path):
                raise UserError(_("File path '%s' does not exist on server.") % self.file_path)
            with open(self.file_path, 'rb') as f:
                return f.read()
        elif self.source_type == 'url':
            if not self.file_url:
                raise UserError(_("Please provide a remote file URL."))
            req = urllib.request.Request(self.file_url, headers={'User-Agent': 'Odoo-ETL-Studio/19.0'})
            try:
                with urllib.request.urlopen(req, timeout=30) as response:
                    return response.read()
            except Exception as e:
                raise UserError(_("Failed to download file from URL '%s': %s") % (self.file_url, str(e)))
        raise UserError(_("Invalid file source configuration."))

    def _parse_csv(self, limit=None):
        content = self._get_file_content_bytes()
        encoding = self.file_encoding or 'utf-8'
        text = content.decode(encoding, errors='replace')
        text_stream = io.StringIO(text)
        delimiter = self.csv_delimiter or ','
        quotechar = self.csv_quotechar or '"'
        
        if self.csv_has_header:
            reader = csv.DictReader(text_stream, delimiter=delimiter, quotechar=quotechar)
            columns = reader.fieldnames or []
            records = []
            for idx, row in enumerate(reader):
                if limit and idx >= limit:
                    break
                records.append(dict(row))
            return records, list(columns)
        else:
            reader = csv.reader(text_stream, delimiter=delimiter, quotechar=quotechar)
            records = []
            columns = []
            for idx, row in enumerate(reader):
                if not columns:
                    columns = [f"col_{i+1}" for i in range(len(row))]
                if limit and idx >= limit:
                    break
                records.append(dict(zip(columns, row)))
            return records, columns

    def _parse_excel(self, limit=None):
        content = self._get_file_content_bytes()
        file_stream = io.BytesIO(content)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_stream, data_only=True)
            sheet = wb.active
            if self.excel_sheet_name:
                if self.excel_sheet_name in wb.sheetnames:
                    sheet = wb[self.excel_sheet_name]
                elif self.excel_sheet_name.isdigit():
                    sheet_idx = int(self.excel_sheet_name)
                    if sheet_idx < len(wb.sheetnames):
                        sheet = wb[wb.sheetnames[sheet_idx]]

            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return [], []
            columns = [str(col).strip() if col is not None else f"col_{i+1}" for i, col in enumerate(rows[0])]
            records = []
            for r_idx, row in enumerate(rows[1:]):
                if limit and r_idx >= limit:
                    break
                record = {columns[c_idx]: (val if val is not None else '') for c_idx, val in enumerate(row)}
                records.append(record)
            return records, columns
        except ImportError:
            raise UserError(_("Python library 'openpyxl' is required to parse Excel files. Please install it on the server."))

    def _parse_json(self, limit=None):
        content = self._get_file_content_bytes()
        encoding = self.file_encoding or 'utf-8'
        data = json.loads(content.decode(encoding, errors='replace'))
        
        # Navigate through custom JSON key path if specified
        if self.json_data_path:
            keys = self.json_data_path.split('.')
            for k in keys:
                if isinstance(data, dict) and k in data:
                    data = data[k]

        if isinstance(data, dict):
            for k, v in data.items():
                if isinstance(v, list):
                    data = v
                    break

        if not isinstance(data, list):
            raise UserError(_("JSON content must resolve to an array of objects."))

        columns = set()
        for item in data[:50]:
            if isinstance(item, dict):
                columns.update(item.keys())
        columns = sorted(list(columns))
        records = data[:limit] if limit else data
        return records, columns

    def _parse_xml(self, limit=None):
        content = self._get_file_content_bytes()
        import xml.etree.ElementTree as ET
        try:
            root = ET.fromstring(content)
            xpath = self.xml_root_xpath or './/record'
            elements = root.findall(xpath)
            if not elements:
                # Try finding all direct children of root
                elements = list(root)

            records = []
            columns = set()
            for idx, el in enumerate(elements):
                if limit and idx >= limit:
                    break
                row = {}
                for child in el:
                    row[child.tag] = child.text or ''
                    columns.add(child.tag)
                records.append(row)
            return records, sorted(list(columns))
        except Exception as e:
            raise UserError(_("Failed to parse XML file: %s") % str(e))

    # ------------------------------------------------------------
    # 2. FOXPRO DBF PARSER
    # ------------------------------------------------------------

    def _parse_foxpro_dbf(self, limit=None):
        """Read Visual FoxPro DBF table natively or via dbfread."""
        content_bytes = None
        target_path = self.file_path
        if self.source_type == 'upload' or not target_path:
            content_bytes = self._get_file_content_bytes()
        else:
            if os.path.isdir(target_path) and self.dbf_table_name:
                target_path = os.path.join(target_path, self.dbf_table_name)
            if not os.path.exists(target_path):
                raise UserError(_("DBF File path '%s' not found.") % target_path)

        encoding = self.file_encoding or 'cp1252'

        try:
            if content_bytes is None and target_path:
                with open(target_path, 'rb') as f:
                    content_bytes = f.read()
            return self._parse_dbf_bytes_natively(content_bytes, encoding=encoding, limit=limit)
        except Exception as err:
            _logger.info("Native DBF parser failed (%s), falling back to dbfread library...", err)
            try:
                import dbfread
                if target_path and os.path.exists(target_path):
                    table = dbfread.DBF(target_path, encoding=encoding, ignore_missing_memofile=True)
                else:
                    import tempfile
                    with tempfile.NamedTemporaryFile(suffix='.dbf', delete=False) as tmp:
                        tmp.write(content_bytes)
                        tmp_path = tmp.name
                    table = dbfread.DBF(tmp_path, encoding=encoding, ignore_missing_memofile=True)
                columns = table.field_names
                records = []
                for idx, record in enumerate(table):
                    if limit and idx >= limit:
                        break
                    records.append(dict(record))
                return records, columns
            except ImportError:
                raise UserError(_("DBF Parsing error: %s") % str(err))

    def _parse_dbf_bytes_natively(self, data, encoding='cp1252', limit=None):
        if len(data) < 32:
            raise UserError(_("Invalid DBF file header: file too small."))
        version, yy, mm, dd, num_records, header_len, record_len = struct.unpack('<BBBBIHH', data[:12])

        fields_data = data[32:header_len]
        fields = []
        offset = 0
        while offset < len(fields_data):
            if fields_data[offset] == 0x0D:
                break
            field_desc = fields_data[offset:offset+32]
            if len(field_desc) < 32:
                break
            field_name_raw = field_desc[:11].replace(b'\x00', b'').decode(encoding, errors='ignore').strip()
            field_type = chr(field_desc[11])
            field_len = field_desc[16]
            field_dec = field_desc[17]
            fields.append({
                'name': field_name_raw,
                'type': field_type,
                'len': field_len,
                'dec': field_dec,
            })
            offset += 32

        columns = [f['name'] for f in fields]
        records = []
        rec_start = header_len

        for rec_idx in range(num_records):
            if limit and rec_idx >= limit:
                break
            rec_offset = rec_start + (rec_idx * record_len)
            if rec_offset + record_len > len(data):
                break
            rec_bytes = data[rec_offset:rec_offset+record_len]
            if rec_bytes and rec_bytes[0:1] == b'*':
                continue
            
            row = {}
            pos = 1
            for f in fields:
                val_bytes = rec_bytes[pos:pos+f['len']]
                pos += f['len']
                val_str = val_bytes.decode(encoding, errors='replace').strip()
                if f['type'] in ('N', 'F'):
                    if val_str:
                        val = float(val_str) if '.' in val_str else int(val_str)
                    else:
                        val = 0
                elif f['type'] == 'L':
                    val = val_str in ('T', 't', 'Y', 'y', '1')
                else:
                    val = val_str
                row[f['name']] = val
            records.append(row)

        return records, columns

    # ------------------------------------------------------------
    # 3. SQL DATABASE & ODBC
    # ------------------------------------------------------------

    def _parse_odbc(self, limit=None):
        if not self.odbc_connection_string:
            raise UserError(_("Please provide an ODBC Connection String."))
        try:
            import pyodbc
            conn = pyodbc.connect(self.odbc_connection_string)
            cursor = conn.cursor()
            query = self.db_query or "SELECT * FROM source_table"
            if limit:
                query = f"SELECT TOP {limit} * FROM ({query}) AS subquery"
            cursor.execute(query)
            columns = [column[0] for column in cursor.description]
            records = []
            for row in cursor.fetchall():
                records.append(dict(zip(columns, row)))
            cursor.close()
            conn.close()
            return records, columns
        except ImportError:
            raise UserError(_("Python library 'pyodbc' is required for ODBC connections. Please install pyodbc."))
        except Exception as e:
            raise UserError(_("ODBC error: %s") % str(e))

    def _parse_sql_db(self, limit=None):
        db_type = self.db_type
        host = self.db_host or 'localhost'
        port = self.db_port or 5432
        dbname = self.db_name
        user = self.db_user
        pwd = self.db_password
        query = self.db_query or "SELECT * FROM source_table"

        if db_type == 'postgresql':
            try:
                import psycopg2
                import psycopg2.extras
                ssl_mode = self.db_ssl_mode or 'prefer'
                conn = psycopg2.connect(host=host, port=port, dbname=dbname, user=user, password=pwd, sslmode=ssl_mode)
                cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                if limit:
                    query = f"SELECT * FROM ({query}) sub LIMIT {limit}"
                cursor.execute(query)
                records = [dict(r) for r in cursor.fetchall()]
                columns = list(records[0].keys()) if records else []
                cursor.close()
                conn.close()
                return records, columns
            except ImportError:
                raise UserError(_("Python library 'psycopg2' is required for PostgreSQL connections."))
            except Exception as e:
                raise UserError(_("PostgreSQL error: %s") % str(e))

        elif db_type == 'mysql':
            try:
                import pymysql
                import pymysql.cursors
                conn = pymysql.connect(
                    host=host, port=port or 3306, user=user, password=pwd, database=dbname,
                    cursorclass=pymysql.cursors.DictCursor
                )
                with conn.cursor() as cursor:
                    if limit:
                        query = f"SELECT * FROM ({query}) sub LIMIT {limit}"
                    cursor.execute(query)
                    records = cursor.fetchall()
                    columns = list(records[0].keys()) if records else []
                    return list(records), columns
            except ImportError:
                raise UserError(_("Python library 'pymysql' is required for MySQL connections."))
            except Exception as e:
                raise UserError(_("MySQL connection error: %s") % str(e))

        elif db_type == 'sqlite':
            try:
                import sqlite3
                db_file = self.db_name or self.file_path
                if not db_file or not os.path.exists(db_file):
                    raise UserError(_("SQLite database file '%s' not found.") % db_file)
                conn = sqlite3.connect(db_file)
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                if limit:
                    query = f"SELECT * FROM ({query}) LIMIT {limit}"
                cursor.execute(query)
                rows = cursor.fetchall()
                records = [dict(r) for r in rows]
                columns = list(records[0].keys()) if records else []
                cursor.close()
                conn.close()
                return records, columns
            except Exception as e:
                raise UserError(_("SQLite error: %s") % str(e))

        elif db_type == 'mssql':
            try:
                import pymssql
                conn = pymssql.connect(server=host, port=port or 1433, user=user, password=pwd, database=dbname, as_dict=True)
                cursor = conn.cursor()
                if limit:
                    query = f"SELECT TOP {limit} * FROM ({query}) AS sub"
                cursor.execute(query)
                records = cursor.fetchall()
                columns = list(records[0].keys()) if records else []
                conn.close()
                return records, columns
            except ImportError:
                raise UserError(_("Python library 'pymssql' or 'pyodbc' is required for SQL Server connections."))
            except Exception as e:
                raise UserError(_("MSSQL connection error: %s") % str(e))

        else:
            raise UserError(_("Database driver for '%s' is not supported.") % db_type)

    # ------------------------------------------------------------
    # 4. REST API & GRAPHQL PARSER
    # ------------------------------------------------------------

    def _parse_rest_api(self, limit=None):
        if not self.api_url:
            raise UserError(_("Please provide a REST API endpoint URL."))
        
        headers = {'Accept': 'application/json', 'Content-Type': 'application/json'}
        if self.api_headers_json:
            try:
                custom_headers = json.loads(self.api_headers_json)
                headers.update(custom_headers)
            except Exception:
                pass

        if self.api_auth_type == 'bearer' and self.api_token:
            headers['Authorization'] = f"Bearer {self.api_token}"
        elif self.api_auth_type == 'api_key' and self.api_token:
            headers[self.api_key_header or 'X-API-Key'] = self.api_token
        elif self.api_auth_type == 'basic' and self.api_user:
            auth_str = f"{self.api_user}:{self.api_password or ''}"
            b64_auth = base64.b64encode(auth_str.encode('utf-8')).decode('utf-8')
            headers['Authorization'] = f"Basic {b64_auth}"

        records = []
        page = 1
        max_pages = 1 if self.api_pagination_type == 'none' else (self.api_max_pages or 20)

        while page <= max_pages:
            url = self.api_url
            params = {}
            if self.api_params_json:
                try:
                    params.update(json.loads(self.api_params_json))
                except Exception:
                    pass

            if self.api_pagination_type == 'page_number':
                params[self.api_page_param or 'page'] = page
                params[self.api_limit_param or 'limit'] = self.api_page_size or 100
            elif self.api_pagination_type == 'offset_limit':
                params[self.api_page_param or 'offset'] = (page - 1) * (self.api_page_size or 100)
                params[self.api_limit_param or 'limit'] = self.api_page_size or 100

            if params:
                query_string = urllib.parse.urlencode(params)
                url = f"{url}?{query_string}" if '?' not in url else f"{url}&{query_string}"

            body_data = None
            if self.api_http_method == 'POST' and self.api_request_body:
                body_data = self.api_request_body.encode('utf-8')

            req = urllib.request.Request(url, data=body_data, headers=headers, method=self.api_http_method or 'GET')
            try:
                with urllib.request.urlopen(req, timeout=30) as response:
                    res_body = json.loads(response.read().decode('utf-8'))
                    page_items = res_body
                    if self.json_data_path:
                        for k in self.json_data_path.split('.'):
                            if isinstance(page_items, dict) and k in page_items:
                                page_items = page_items[k]

                    if isinstance(page_items, dict):
                        for k, v in page_items.items():
                            if isinstance(v, list):
                                page_items = v
                                break

                    if not isinstance(page_items, list):
                        page_items = [page_items] if isinstance(page_items, dict) else []

                    if not page_items:
                        break

                    records.extend(page_items)
                    if limit and len(records) >= limit:
                        records = records[:limit]
                        break

            except urllib.error.HTTPError as e:
                err_text = e.read().decode('utf-8', errors='replace')
                raise UserError(_("REST API HTTP Error %s: %s") % (e.code, err_text))
            except Exception as e:
                raise UserError(_("REST API Error: %s") % str(e))

            if self.api_pagination_type == 'none':
                break
            page += 1

        columns = set()
        for item in records[:50]:
            if isinstance(item, dict):
                columns.update(item.keys())
        return records, sorted(list(columns))

    def _parse_graphql(self, limit=None):
        if not self.api_url:
            raise UserError(_("Please provide a GraphQL endpoint URL."))
        query = self.api_request_body
        if not query:
            raise UserError(_("Please provide a GraphQL query payload."))

        headers = {'Content-Type': 'application/json'}
        if self.api_token:
            headers['Authorization'] = f"Bearer {self.api_token}"

        payload = {'query': query}
        req = urllib.request.Request(self.api_url, data=json.dumps(payload).encode('utf-8'), headers=headers, method='POST')
        try:
            with urllib.request.urlopen(req, timeout=30) as response:
                body = json.loads(response.read().decode('utf-8'))
                data = body.get('data', {})
                # Extract first list found in data
                records = []
                for k, v in data.items():
                    if isinstance(v, list):
                        records = v
                        break
                    elif isinstance(v, dict):
                        for sub_k, sub_v in v.items():
                            if isinstance(sub_v, list):
                                records = sub_v
                                break
                if limit:
                    records = records[:limit]
                columns = sorted(list(records[0].keys())) if records else []
                return records, columns
        except Exception as e:
            raise UserError(_("GraphQL error: %s") % str(e))

    # ------------------------------------------------------------
    # 5. CLOUD STORAGE & GOOGLE SHEETS
    # ------------------------------------------------------------

    def _parse_cloud_storage(self, limit=None):
        if self.cloud_provider == 's3':
            try:
                import boto3
                s3_kwargs = {
                    'aws_access_key_id': self.cloud_access_key,
                    'aws_secret_access_key': self.cloud_secret_key,
                    'region_name': self.cloud_region or 'us-east-1',
                }
                if self.cloud_endpoint_url:
                    s3_kwargs['endpoint_url'] = self.cloud_endpoint_url

                s3 = boto3.client('s3', **s3_kwargs)
                response = s3.get_object(Bucket=self.cloud_bucket, Key=self.cloud_object_key)
                content = response['Body'].read()
                
                # Parse based on file extension
                key = (self.cloud_object_key or '').lower()
                if key.endswith('.json'):
                    data = json.loads(content.decode(self.file_encoding or 'utf-8', errors='replace'))
                    records = data if isinstance(data, list) else [data]
                    cols = sorted(list(records[0].keys())) if records else []
                    return records[:limit] if limit else records, cols
                else:
                    # CSV default
                    text = content.decode(self.file_encoding or 'utf-8', errors='replace')
                    reader = csv.DictReader(io.StringIO(text), delimiter=self.csv_delimiter or ',')
                    records = [dict(r) for idx, r in enumerate(reader) if not limit or idx < limit]
                    return records, list(reader.fieldnames or [])
            except ImportError:
                raise UserError(_("Python library 'boto3' is required for AWS S3 cloud storage."))
            except Exception as e:
                raise UserError(_("AWS S3 error: %s") % str(e))

        elif self.cloud_provider == 'sftp':
            try:
                import paramiko
                ssh = paramiko.SSHClient()
                ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                ssh.connect(
                    hostname=self.sftp_host, port=self.sftp_port or 22,
                    username=self.sftp_user, password=self.sftp_password, timeout=15
                )
                sftp = ssh.open_sftp()
                with sftp.open(self.cloud_object_key, 'rb') as f:
                    content = f.read()
                sftp.close()
                ssh.close()
                text = content.decode(self.file_encoding or 'utf-8', errors='replace')
                reader = csv.DictReader(io.StringIO(text), delimiter=self.csv_delimiter or ',')
                records = [dict(r) for idx, r in enumerate(reader) if not limit or idx < limit]
                return records, list(reader.fieldnames or [])
            except ImportError:
                raise UserError(_("Python library 'paramiko' is required for SFTP storage."))
            except Exception as e:
                raise UserError(_("SFTP error: %s") % str(e))
        else:
            raise UserError(_("Cloud provider '%s' is not implemented.") % self.cloud_provider)

    def _parse_google_sheets(self, limit=None):
        url = self.sheets_url
        if not url:
            raise UserError(_("Please provide a Google Sheets URL."))
        # Convert standard URL to public CSV export URL
        sheet_id_match = re.search(r'/d/([a-zA-Z0-9-_]+)', url)
        if not sheet_id_match:
            raise UserError(_("Invalid Google Sheet URL format. Expected URL containing /d/SHEET_ID/"))
        sheet_id = sheet_id_match.group(1)
        gid = self.sheets_gid or '0'
        export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"

        req = urllib.request.Request(export_url, headers={'User-Agent': 'Odoo-ETL-Studio/19.0'})
        try:
            with urllib.request.urlopen(req, timeout=30) as response:
                content = response.read().decode('utf-8', errors='replace')
                reader = csv.DictReader(io.StringIO(content))
                records = []
                for idx, row in enumerate(reader):
                    if limit and idx >= limit:
                        break
                    records.append(dict(row))
                return records, list(reader.fieldnames or [])
        except Exception as e:
            raise UserError(_("Google Sheets export failed. Ensure sheet sharing is set to 'Anyone with link can view': %s") % str(e))

    # ------------------------------------------------------------
    # 6. ODOO RPC PARSER
    # ------------------------------------------------------------

    def _parse_odoo_rpc(self, limit=None):
        import xmlrpc.client
        url = (self.rpc_url or '').rstrip('/')
        db = self.rpc_db
        user = self.rpc_user
        password = self.rpc_password
        model = self.rpc_model
        domain = eval(self.rpc_domain or '[]')

        common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common')
        uid = common.authenticate(db, user, password, {})
        if not uid:
            raise UserError(_("Remote Odoo authentication failed."))
        models_api = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object')
        ids = models_api.execute_kw(db, uid, password, model, 'search', [domain], {'limit': limit or 100})
        fields_info = models_api.execute_kw(db, uid, password, model, 'fields_get', [], {'attributes': ['string', 'type']})
        columns = sorted(list(fields_info.keys()))
        records = models_api.execute_kw(db, uid, password, model, 'read', [ids], {'fields': columns})
        return records, columns
